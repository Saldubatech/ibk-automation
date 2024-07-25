import datetime
from typing import Optional

import pandas as pd
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from salduba.common.configuration import Defaults
from salduba.corvino.io.parse_input import InputRow
from salduba.corvino.persistence.movement_record import MovementRecord2
from salduba.ib_tws_proxy.contracts.contract_repo import ContractRecord2
from salduba.ib_tws_proxy.operations import ErrorResponse

contract_columns = [
  'expires_on',
  'con_id',
  'symbol',
  'sec_type',
  'lookup_exchange',
  'exchange',
  'primary_exchange',
  'currency',
  'local_symbol',
  'trading_class',
  'sec_id_type',
  'sec_id',
  'rid',
  'at'
]


input_columns = ['ticker', 'trade', 'symbol', 'country', 'raw_type', 'ibk_type', 'currency', 'exchange', 'exchange2']


error_columns = ['Error', 'opId', 'errorCode', 'errorString']


movement_columns = [
  "status",
  'batch',
  'ticker',
  'trade',
  # 'nombre',
  'symbol',
  'raw_type',
  'ibk_type',
  'country',
  'currency',
  'exchange',
  'exchange2',
  "rid",
  "at"
]


class ResultsBatch:
  def __init__(self,
               atTime: datetime.datetime,
               message: str,
               inputs: list[InputRow],
               known: list[ContractRecord2],
               updated: list[ContractRecord2],
               unknown: list[InputRow],
               movements_placed: list[MovementRecord2],
               errors: dict[str, list[ErrorResponse]]) -> None:
    self.atTime = atTime
    self.message = message
    self.errors: dict[str, list[ErrorResponse]] = errors
    self.inputs = inputs
    self.known = known
    self.movements = movements_placed
    self.updated = updated
    self.unknown = unknown
    self.input_sheet: str = Defaults.output.inputs_sheet
    self.known_sheet: str = Defaults.output.known_sheet
    self.updated_sheet: str = Defaults.output.updated_sheet
    self.missing_sheet: str = Defaults.output.missing_sheet
    self.movement_sheet: str = Defaults.output.movement_sheet
    self.error_sheet: str = Defaults.output.error_sheet
    self.file_prefix: str = Defaults.output.file_prefix
    self.filename = Defaults.output.file_name

  @staticmethod
  def _contracts_pd(contract_records: list[ContractRecord2]) -> pd.DataFrame:
    return pd.DataFrame(
      data=[{c:  cr.__dict__[c] for c in contract_columns} for cr in contract_records],
      columns=contract_columns
    )

  @staticmethod
  def _input_pd(input_rows: list[InputRow]) -> pd.DataFrame:
    return pd.DataFrame(
      columns=input_columns,
      data=[{c: ir.__dict__[c] for c in input_columns} for ir in input_rows]
    )

  @staticmethod
  def _movements_pd(movements: list[MovementRecord2]) -> pd.DataFrame:
    return pd.DataFrame(
      columns=movement_columns,
      data=[{c: m.__dict__[c] for c in movement_columns} for m in movements]
    )

  @staticmethod
  def _write_one(wb: Workbook, data: Optional[pd.DataFrame], columns: list[str], sheet_name: str) -> None:
    if data is not None:
      sheet: Worksheet = wb.create_sheet(sheet_name)
      # Ensure the sheet is clean.
      sheet.delete_rows(1, sheet.max_row)
      # Ensure the df is well indexed
      data.reset_index()
      for idx, c_header in enumerate(columns):
        sheet.cell(row=1, column=idx+1).value = c_header
      for r_idx, r in data.iterrows():  # pyright: ignore
        for c_idx, c_name in enumerate(columns):
          sheet.cell(row=int(r_idx)+2, column=c_idx+1).value = r[c_name]  # type: ignore

  def write_xlsx(self, override_filename: Optional[str] = None) -> None:
    wb: Workbook = Workbook()

    ResultsBatch._write_one(
      wb,
      ResultsBatch._input_pd(self.inputs) if self.inputs else None,
      input_columns,
      self.input_sheet)

    ResultsBatch._write_one(
      wb,
      ResultsBatch._contracts_pd(self.known) if self.known else None,
      contract_columns,
      self.known_sheet)

    ResultsBatch._write_one(
      wb,
      ResultsBatch._contracts_pd(self.updated) if self.updated else None,
      contract_columns,
      self.updated_sheet)

    ResultsBatch._write_one(
      wb,
      ResultsBatch._input_pd(self.unknown) if self.unknown else None,
      input_columns,
      self.missing_sheet)

    ResultsBatch._write_one(
      wb,
      ResultsBatch._movements_pd(self.movements) if self.movements else None,
      movement_columns,
      self.movement_sheet)

    if self.errors:
      error_sheet: Worksheet = wb.create_sheet(self.error_sheet)
      for c_idx, e_tag in enumerate(error_columns):
        error_sheet.cell(1, c_idx+1).value = e_tag
      current_row = 2
      for k, l_err in self.errors.items():
        error_sheet.cell(row=current_row, column=1).value = k
        for err in l_err:
          error_sheet.cell(row=current_row, column=2).value = str(err.errorCode)
          error_sheet.cell(row=current_row, column=3).value = str(err.errorString)
          current_row += 1
    wb.remove(wb['Sheet'])
    if override_filename:
      wb.save(override_filename)
    else:
      wb.save(self.filename)
