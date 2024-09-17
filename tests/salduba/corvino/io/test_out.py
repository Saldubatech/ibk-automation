import datetime
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from salduba.corvino.io.parse_input import InputRow
from salduba.corvino.io.results_out import ResultsBatch, contract_columns, error_columns, input_columns, movement_columns
from salduba.corvino.persistence.movement_record import MovementRecord2, MovementStatus
from salduba.ib_tws_proxy.contracts.contract_repo import ContractRecord2
from salduba.ib_tws_proxy.domain.enumerations import Country, Currency, Exchange, SecType
from salduba.ib_tws_proxy.operations import ErrorResponse
from salduba.ib_tws_proxy.orders.OrderRepo import OrderRecord2
from salduba.util.logging import init_logging
from salduba.util.tests import findTestsRoot

_maybeTr = findTestsRoot()
_tr = _maybeTr if _maybeTr else "./"
init_logging(Path(_tr, "resources/logging.yaml"))
_logger = logging.getLogger(__name__)


def orderProbe(seed: int) -> OrderRecord2:
  return OrderRecord2(
    rid=f'{seed}_rid',
    at=seed*11,
    orderId=f"{seed}_orderId",
    transmit=False,
  )


def contractProbe(seed: int) -> ContractRecord2:
  return ContractRecord2(
    rid=f'{seed}_rid',
    at=seed*11,
    expires_on=seed*22,
    con_id=seed*101,
    symbol=f"{seed}Symbol",
    sec_type=SecType.STK,
    last_trade_date_or_contract_month=None,
    strike=float(seed)*1.11111,
    right=f"{seed}_right",
    multiplier=f"{seed}_multiplier",
    lookup_exchange=Exchange.NYSE,
    exchange=Exchange.ISLAND,
    primary_exchange=Exchange.NYSE,
    currency=Currency.USD,
    local_symbol=f"{seed}_Local",
    trading_class=f"{seed}_TradingClass",
    sec_id_type=f"{seed}_secIdType",
    sec_id=f"{seed}_SecId",
    combo_legs_description=f'{seed}combo_legs_desc',
    include_expired=False,
    delta_neutral_contract=None  # f"{seed}dnc_fk",
  )


def inputProbe(seed: int) -> InputRow:
  return InputRow(
    f"{seed}_ticker",
    seed*1010,
    f"{seed}_symbol",
    Country.US,
    f"{seed}_raw_type",
    SecType.STK,
    Currency.JPY,
    Exchange.IBIS,
    Exchange.NYSE
  )


def movementProbe(batch: str, seed: int) -> MovementRecord2:
  return MovementRecord2(
    rid=f'{seed}_rid',
    at=seed*11,
    status=MovementStatus.CONFIRMED,
    batch=batch,
    ticker=f"{seed}_ticker",
    trade=seed*100,
    nombre=f"{seed}_nombre",
    symbol=f"{seed}_symbol",
    raw_type=f"{seed}_raw_type",
    ibk_type=SecType.STK,
    country=Country.US,
    currency=Currency.USD,
    exchange=Exchange.ISLAND,
    exchange2=Exchange.NYSE,
    contract=contractProbe(seed),  # f"{seed}_contract_fk",
    order=orderProbe(seed)  # f"{seed}_order_fk"
  )


sample_contracts = [contractProbe(i) for i in range(1, 12)]

sample_inputs = [inputProbe(i) for i in range(1, 7)]


sample_errors = {
  f"error{en}":
    [ErrorResponse(i, i*10, f"{i}_error_str", f"{i}_advanced_error_json")
     for i in range(1, 3)]
  for en in range(1, 5)
}


sample_movements = [movementProbe("TestBatch", i) for i in range(1, 7)]


def check_sheet(
    sheets: list[str], columns: list[str], probes: list[dict[str, Any]], sheet: Worksheet, date_compare: bool = False) -> None:
  assert sheet
  assert sheet.max_row == len(probes)+1
  for c_idx, h in enumerate(columns):
    assert sheet.cell(row=1, column=c_idx+1).value == h
    for r_idx, ir in enumerate(probes):
      cell_value = sheet.cell(row=r_idx+2, column=c_idx+1).value
      if h in ['expires_on', 'at'] and date_compare:
        assert cell_value == datetime.datetime.fromtimestamp(ir[h]/1e3).strftime("%Y-%m-%d %H:%M:%S"), \
          f"Checking {h} in row {r_idx+2} with value {cell_value} against {ir[h]}"
      else:
        assert cell_value == ir[h], f"Checking {h} in row {r_idx+2} with value {cell_value} against {ir[h]}"


def test_input() -> None:
  probe = ResultsBatch(
    datetime.datetime.now(),
    "Testing InputRows render",
    sample_inputs,
    [],
    [],
    [],
    [],
    {}
  )
  probe.write_xlsx()
  result = load_workbook(probe.filename)
  assert len(result.sheetnames) == 1
  assert result.sheetnames[0] == probe.input_sheet
  input_sheet = result[probe.input_sheet]
  check_sheet([probe.input_sheet], input_columns, [r.__dict__ for r in probe.inputs], input_sheet)


def test_missing() -> None:
  probe = ResultsBatch(
    datetime.datetime.now(),
    "Testing InputRows render",
    [],
    [],
    [],
    sample_inputs,
    [],
    {}
  )
  probe.write_xlsx()
  result = load_workbook(probe.filename)
  assert len(result.sheetnames) == 1
  assert result.sheetnames[0] == probe.missing_sheet
  sheet = result[probe.missing_sheet]
  check_sheet([probe.input_sheet], input_columns, [r.__dict__ for r in probe.unknown], sheet)


def test_known() -> None:
  probe = ResultsBatch(
    atTime=datetime.datetime.now(),
    message="Testing Knowns render",
    inputs=[],
    known=sample_contracts,
    updated=[],
    unknown=[],
    movements_placed=[],
    errors={}
  )
  probe.write_xlsx()
  result = load_workbook(probe.filename)
  assert len(result.sheetnames) == 1
  assert result.sheetnames[0] == probe.known_sheet
  sheet = result[probe.known_sheet]
  check_sheet([probe.known_sheet], contract_columns, [r.__dict__ for r in probe.known], sheet, True)


def test_updated() -> None:
  probe = ResultsBatch(
    datetime.datetime.now(),
    "Testing Knowns render",
    [],
    [],
    sample_contracts,
    [],
    [],
    {}
  )
  probe.write_xlsx()
  result = load_workbook(probe.filename)
  assert len(result.sheetnames) == 1
  assert result.sheetnames[0] == probe.updated_sheet
  sheet = result[probe.updated_sheet]
  check_sheet([probe.updated_sheet], contract_columns, [r.__dict__ for r in probe.updated], sheet, True)


def test_orders() -> None:
  probe = ResultsBatch(
    datetime.datetime.now(),
    "Testing Knowns render",
    [],
    [],
    [],
    [],
    sample_movements,
    {}
  )
  probe.write_xlsx()
  result = load_workbook(probe.filename)
  assert len(result.sheetnames) == 1
  assert result.sheetnames[0] == probe.movement_sheet
  sheet = result[probe.movement_sheet]
  check_sheet([probe.movement_sheet], movement_columns, [r.__dict__ for r in probe.movements], sheet)


def test_errors() -> None:
  probe = ResultsBatch(
    datetime.datetime.now(),
    "Testing Knowns render",
    [],
    [],
    [],
    [],
    [],
    sample_errors
  )
  probe.write_xlsx()
  result = load_workbook(probe.filename)
  assert len(result.sheetnames) == 1
  assert result.sheetnames[0] == probe.error_sheet
  sheet = result[probe.error_sheet]
  for c_idx, h in enumerate(error_columns):
    assert sheet.cell(row=1, column=c_idx+1).value == h

  current_row = 2
  error_row = 1
  for k, errs in probe.errors.items():
    for err in errs:
      if error_row == 1:
        assert sheet.cell(current_row, 1).value == k
      else:
        assert sheet.cell(current_row, 1).value is None
      assert sheet.cell(current_row, 2).value == str(err.errorCode)
      assert sheet.cell(current_row, 3).value == str(err.errorString)
      current_row += 1
      error_row += 1
    error_row = 1
  assert sheet.max_row == current_row - 1


def test_all_together() -> None:
  probe = ResultsBatch(
    datetime.datetime.now(),
    "Testing All render",
    sample_inputs,
    sample_contracts,
    sample_contracts,
    sample_inputs,
    sample_movements,
    sample_errors
  )
  probe.write_xlsx()
  result = load_workbook(probe.filename)
  assert len(result.sheetnames) == 6
  assert result.sheetnames == [
    probe.input_sheet,
    probe.known_sheet,
    probe.updated_sheet,
    probe.missing_sheet,
    probe.movement_sheet,
    probe.error_sheet]

  sheet = result[probe.input_sheet]
  check_sheet([probe.input_sheet], input_columns, [r.__dict__ for r in probe.inputs], sheet)

  sheet = result[probe.known_sheet]
  check_sheet([probe.known_sheet], contract_columns, [r.__dict__ for r in probe.known], sheet, True)

  sheet = result[probe.updated_sheet]
  check_sheet([probe.updated_sheet], contract_columns, [r.__dict__ for r in probe.updated], sheet, True)

  sheet = result[probe.missing_sheet]
  check_sheet([probe.missing_sheet], input_columns, [r.__dict__ for r in probe.unknown], sheet)

  sheet = result[probe.movement_sheet]
  check_sheet([probe.movement_sheet], movement_columns, [r.__dict__ for r in probe.movements], sheet)

  sheet = result[probe.error_sheet]
  for c_idx, h in enumerate(error_columns):
    assert sheet.cell(row=1, column=c_idx+1).value == h

  current_row = 2
  error_row = 1
  for k, errs in probe.errors.items():
    for err in errs:
      if error_row == 1:
        assert sheet.cell(current_row, 1).value == k
      else:
        assert sheet.cell(current_row, 1).value is None
      assert sheet.cell(current_row, 2).value == str(err.errorCode)
      assert sheet.cell(current_row, 3).value == str(err.errorString)
      current_row += 1
      error_row += 1
    error_row = 1
  assert sheet.max_row == current_row - 1
